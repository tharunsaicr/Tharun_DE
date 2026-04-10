# Databricks notebook source
from openpyxl import load_workbook

# COMMAND ----------

dbutils.widgets.text(name = "ProcessName", defaultValue = "")
ProcessName = dbutils.widgets.get("ProcessName")

dbutils.widgets.text(name = "IpPath", defaultValue = "")
IpPath = dbutils.widgets.get("IpPath")

dbutils.widgets.text(name = "DestinationPath", defaultValue = "")
DestinationPath = dbutils.widgets.get("DestinationPath")

dbutils.widgets.text(name = "FileName", defaultValue = "")
FileName = dbutils.widgets.get("FileName")

print(IpPath)
print(ProcessName)
print(FileName)
print(DestinationPath)

# COMMAND ----------

# MAGIC %run /kgsfinance/common_utilities/install_mssql

# COMMAND ----------

# MAGIC %run /kgsfinance/common_utilities/connection_configuration

# COMMAND ----------

# MAGIC %run /kgsfinance/common_utilities/common_components

# COMMAND ----------

workbook=load_workbook(filename="/dbfs/mnt"+str(IpPath)+str(FileName),read_only=True)
sheet_names_1=workbook.sheetnames

# COMMAND ----------

report_name=IpPath.split("/")[-5]
print(report_name)

# COMMAND ----------

print(sheet_names_1)

# COMMAND ----------

if report_name == 'functional_report':
    FileName=FileName[:-12]
    print(FileName)

# COMMAND ----------

query= "SELECT DISTINCT(Sheet_Name) from dbo.finance_config_table where Process_Name='{}' and File_Name = '{}'".format(ProcessName,FileName)

# COMMAND ----------

print(query)

# COMMAND ----------

df = spark.read\
  .format("jdbc")\
  .option("driver","com.microsoft.sqlserver.jdbc.SQLServerDriver")\
  .option("url", jdbcUrl)\
  .option("query",query)\
  .option("user", username)\
  .option("password", password)\
  .load()

# COMMAND ----------

display(df)

# COMMAND ----------

config_sheet_names=df.select("Sheet_Name").rdd.flatMap(lambda x:x).collect()

# COMMAND ----------

print(config_sheet_names)

# COMMAND ----------

final_list=[]
for i in sheet_names_1:
    if i in config_sheet_names:
        final_list.append(i)
    

# COMMAND ----------

print(final_list)

# COMMAND ----------

df.count()

# COMMAND ----------

# if len(final_list) != df.count():
#     raise Exception("The {} Process {} workbook does not contain all the sheets".format(ProcessName,FileName))

# COMMAND ----------

dbutils.notebook.exit(final_list)

# COMMAND ----------

